"""
Conciliacion bancaria asistida.

Compara el libro contable de la empresa contra el extracto del banco y
clasifica cada movimiento, exactamente como lo haria una persona a mano
pero en segundos en lugar de horas.

Uso:
    python conciliar.py

Genera output/conciliacion_resultado.xlsx con cada fila coloreada segun
su estado:

    Verde    -> Conciliado (coincide en ambos lados)
    Amarillo -> Pendiente en banco (esta en libros, el banco aun no lo refleja)
    Naranja  -> Pendiente en libros (esta en el banco, nadie lo registro)
    Rojo     -> Revisar diferencia / posible duplicado

No usa ningun servicio externo: el "emparejamiento inteligente" se hace
comparando fecha, valor y texto de la descripcion con la libreria
estandar de Python (difflib). La misma logica es la que hacen por debajo
las herramientas de conciliacion asistidas por IA; con un modelo de
lenguaje (Claude, ChatGPT, etc.) el emparejamiento por texto es todavia
mas robusto porque entiende sinonimos y abreviaturas del banco.
"""

from __future__ import annotations

import difflib
import re
import unicodedata
from dataclasses import dataclass, field

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

FECHA_TOLERANCIA_DIAS = 3
VALOR_TOLERANCIA_RELATIVA = 0.08  # 8% -> suficiente para detectar un digito mal tipeado
SIMILITUD_MINIMA_EXACTA = 0.25
SIMILITUD_MINIMA_DIFERENCIA = 0.45

COLORES = {
    "Conciliado": "C6EFCE",
    "Pendiente en banco": "FFEB9C",
    "Pendiente en libros": "FFD9B3",
    "Revisar diferencia": "FFC7CE",
    "Posible duplicado": "FFC7CE",
}


def normalizar(texto: str) -> str:
    texto = unicodedata.normalize("NFKD", str(texto)).encode("ascii", "ignore").decode()
    texto = re.sub(r"[^a-zA-Z0-9 ]", " ", texto).lower()
    return re.sub(r"\s+", " ", texto).strip()


def similitud(a: str, b: str) -> float:
    return difflib.SequenceMatcher(None, normalizar(a), normalizar(b)).ratio()


@dataclass
class Movimiento:
    origen: str
    idx: int
    fecha: pd.Timestamp
    descripcion: str
    valor: float
    estado: str = ""
    pareja_idx: int | None = None
    detalle: str = ""


def cargar(ruta: str, origen: str) -> list[Movimiento]:
    df = pd.read_excel(ruta)
    df["Fecha"] = pd.to_datetime(df["Fecha"])
    return [
        Movimiento(origen=origen, idx=i, fecha=row.Fecha, descripcion=row.Descripcion, valor=float(row.Valor))
        for i, row in enumerate(df.itertuples())
    ]


def conciliar(libro: list[Movimiento], banco: list[Movimiento]) -> list[Movimiento]:
    banco_usado = [False] * len(banco)

    def candidatos(mov: Movimiento):
        for j, b in enumerate(banco):
            if banco_usado[j]:
                continue
            if abs((b.fecha - mov.fecha).days) <= FECHA_TOLERANCIA_DIAS:
                yield j, b

    # 1) buscar coincidencia exacta de valor (mismo signo, mismo monto) con la mejor descripcion
    for mov in libro:
        mejor = None
        for j, b in candidatos(mov):
            if b.valor == mov.valor:
                score = similitud(mov.descripcion, b.descripcion)
                if score >= SIMILITUD_MINIMA_EXACTA and (mejor is None or score > mejor[1]):
                    mejor = (j, score)
        if mejor:
            j, _ = mejor
            banco_usado[j] = True
            mov.estado = "Conciliado"
            mov.pareja_idx = banco[j].idx
            banco[j].estado = "Conciliado"
            banco[j].pareja_idx = mov.idx

    # 2) para lo que quedo sin pareja, buscar un valor "parecido" (posible error de digitacion)
    for mov in libro:
        if mov.estado:
            continue
        mejor = None
        for j, b in candidatos(mov):
            if b.valor == 0:
                continue
            diff_rel = abs(b.valor - mov.valor) / max(abs(mov.valor), abs(b.valor))
            if diff_rel <= VALOR_TOLERANCIA_RELATIVA:
                score = similitud(mov.descripcion, b.descripcion)
                if score >= SIMILITUD_MINIMA_DIFERENCIA and (mejor is None or score > mejor[1]):
                    mejor = (j, score, b.valor)
        if mejor:
            j, _, valor_banco = mejor
            banco_usado[j] = True
            mov.estado = "Revisar diferencia"
            mov.pareja_idx = banco[j].idx
            mov.detalle = f"Libros: {mov.valor:,.0f}  |  Banco: {valor_banco:,.0f}".replace(",", ".")
            banco[j].estado = "Revisar diferencia"
            banco[j].pareja_idx = mov.idx
            banco[j].detalle = mov.detalle

    # 3) lo que sigue sin pareja en libros: el banco todavia no lo refleja
    for mov in libro:
        if not mov.estado:
            mov.estado = "Pendiente en banco"
            mov.detalle = "Registrado en libros, el banco aun no lo muestra"

    # 4) lo que sobra en el banco: puede ser un movimiento propio del banco (comision,
    #    rendimientos) o un cobro duplicado si ya existe otro movimiento identico conciliado
    for j, b in enumerate(banco):
        if banco_usado[j]:
            continue
        duplicado_de = None
        for otro in banco:
            if otro.idx != b.idx and otro.estado == "Conciliado" and otro.valor == b.valor:
                if similitud(otro.descripcion, b.descripcion) >= 0.6 and abs((otro.fecha - b.fecha).days) <= 1:
                    duplicado_de = otro
                    break
        if duplicado_de is not None:
            b.estado = "Posible duplicado"
            b.detalle = f"Coincide con un movimiento ya conciliado el {duplicado_de.fecha.date()}"
        else:
            b.estado = "Pendiente en libros"
            b.detalle = "Aparece en el banco, no esta registrado en libros"

    return libro + banco


def exportar_excel(movimientos: list[Movimiento], ruta_salida: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Conciliacion"

    encabezados = ["Origen", "Fecha", "Descripcion", "Valor", "Estado", "Detalle"]
    ws.append(encabezados)
    for col in range(1, len(encabezados) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    orden_estado = {
        "Conciliado": 0,
        "Revisar diferencia": 1,
        "Posible duplicado": 1,
        "Pendiente en banco": 2,
        "Pendiente en libros": 3,
    }
    movimientos_ordenados = sorted(movimientos, key=lambda m: (orden_estado.get(m.estado, 9), m.fecha, m.origen))

    for m in movimientos_ordenados:
        ws.append([m.origen, m.fecha.date().isoformat(), m.descripcion, m.valor, m.estado, m.detalle])
        fill = PatternFill(start_color=COLORES[m.estado], end_color=COLORES[m.estado], fill_type="solid")
        for col in range(1, len(encabezados) + 1):
            ws.cell(row=ws.max_row, column=col).fill = fill

    anchos = [12, 12, 42, 14, 20, 46]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    wb.save(ruta_salida)
    print(f"Reporte guardado en {ruta_salida}")


def resumen(movimientos: list[Movimiento]) -> None:
    conteo: dict[str, int] = {}
    for m in movimientos:
        conteo[m.estado] = conteo.get(m.estado, 0) + 1
    total = len(movimientos)
    conciliados = conteo.get("Conciliado", 0)
    print("\n== Resumen de la conciliacion ==")
    for estado, n in sorted(conteo.items(), key=lambda kv: -kv[1]):
        print(f"  {estado:<22} {n}")
    pct = 100 * conciliados / total if total else 0
    print(f"\n{conciliados} de {total} movimientos ({pct:.0f}%) coincidieron automaticamente sin intervencion humana.")


if __name__ == "__main__":
    libro = cargar("data/libro_contable.xlsx", origen="Libro contable")
    banco = cargar("data/extracto_banco.xlsx", origen="Extracto banco")
    resultado = conciliar(libro, banco)
    exportar_excel(resultado, "output/conciliacion_resultado.xlsx")
    resumen(resultado)
